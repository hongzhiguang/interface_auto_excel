from openpyxl import load_workbook
from conf.ProjVar import *
import os
import traceback

class ExcelParse(object):
    """操作表格"""

    def __init__(self,excel_path):
        if not os.path.exists(excel_path):
            self.wb = None
        self.excel_path = excel_path
        self.wb = load_workbook(excel_path)
        self.ws = self.set_sheet_by_name(self.wb.sheetnames[0])

    def set_sheet_by_name(self,name):
        if name in self.wb.sheetnames:
            self.ws = self.wb[name]
            return self.ws
        self.ws = None
        return self.ws

    def get_rows(self,col_no):
        if not isinstance(col_no,int):
            return None
        try:
            return list(self.ws.columns)[col_no-1]
        except:
            traceback.print_exc()

    def get_cols(self,row_no):
        if not isinstance(row_no,int):
            return None
        try:
            return list(self.ws.rows)[row_no-1]
        except:
            traceback.print_exc()

    @property
    def get_min_row(self):
        return self.ws.min_row

    @property
    def get_max_row(self):
        return self.ws.max_row

    @property
    def get_min_col(self):
        return self.ws.min_column

    @property
    def get_max_col(self):
        return self.ws.max_column

    def get_cell_value(self,row_no,col_no):
        if (not isinstance(row_no, int)) or (not isinstance(col_no, int)):
            return None
        try:
            return self.ws.cell(row=row_no, column=col_no).value
        except:
            traceback.print_exc()

    def write_cell(self,row_no,col_no,content):
        if (not isinstance(row_no, int)) or (not isinstance(col_no, int)):
            return None
        try:
            self.ws.cell(row=row_no, column=col_no).value = content
            self._save()
        except:
            traceback.print_exc()

    def _save(self):
        try:
            self.wb.save(self.excel_path)
        except PermissionError as e:
            raise PermissionError("Please close file first.")


if __name__ == "__main__":
    testExcel = ExcelParse(testcase_path)
    testExcel.set_sheet_by_name("注册接口用例")
    print(testExcel.get_rows(API_active_col_no))
    print(testExcel.get_min_row,testExcel.get_max_row)
    print(testExcel.get_min_col,testExcel.get_max_col)